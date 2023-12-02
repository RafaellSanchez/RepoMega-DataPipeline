import openpyxl
import time
import pandas as pd


# Carregue o arquivo Excel
file = '/workspaces/RepoMega-DataPipeline/data/xlsx/Mega-Sena.xlsx'
path = '/workspaces/RepoMega-DataPipeline/data/files/'
wb = openpyxl.load_workbook(filename=file)
file_mg = f'full-data-mega'

name = path + file_mg

dict_dfs = {}

time.sleep(3)
print('Iterando sobre a planilha')
# Iterar sobre as planilhas e criar DataFrames
for sheet in wb.sheetnames:
    sheet_data = []
    for row in wb[sheet].iter_rows(values_only=True):
        sheet_data.append(row)
    
    df = pd.DataFrame(sheet_data[1:], columns=sheet_data[0])
    dict_dfs[sheet] = df
time.sleep(3)
print('DataFrame criado')
# Iterar sobre as planilhas no dicionário e salvar em arquivos de texto delimitados por ponto e vírgula
for sheet_name, df in dict_dfs.items():
    file_name = name + sheet_name + '.txt'
    df.to_csv(file_name, sep=';', index=False)
time.sleep(3)
print('Iterando sobre o df')
print('Salvando o df com delimitador ponto e virgula')
print(f'Nome do arquivo {file_name}')